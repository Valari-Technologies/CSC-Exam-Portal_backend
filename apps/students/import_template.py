"""The student bulk-import column contract, and the downloadable templates for it.

This module is the single definition of which columns an import file may carry. The
importer validates against it and the template endpoint generates from it, so a template
downloaded from the app is by construction a file the importer accepts.
"""
from __future__ import annotations

import csv
import io
from typing import Dict, List, Tuple

REQUIRED_COLUMNS: Tuple[str, ...] = ('full_name', 'roll_number')

# Email is optional: students sign in with a Student ID, so an email is contact
# information for them rather than a credential. Leaving the cell blank is normal.
OPTIONAL_COLUMNS: Tuple[str, ...] = (
    'email',
    'date_of_birth',
    'admission_number',
    'gender',
    'parent_name',
    'parent_phone',
)

# The order columns appear in the template. Email stays near the front, where people
# expect it, even though it is now optional — hence a hand-written order rather than
# required-then-optional.
ALL_COLUMNS: Tuple[str, ...] = (
    'full_name',
    'email',
    'roll_number',
    'date_of_birth',
    'admission_number',
    'gender',
    'parent_name',
    'parent_phone',
)

COLUMN_HELP: Dict[str, str] = {
    'full_name': "Required. The student's full name.",
    'roll_number': 'Required. Unique within the chosen class and section.',
    'email': 'Optional. Leave blank if the student has no email. Must be unique if given.',
    'date_of_birth': 'Optional. Date of birth as YYYY-MM-DD (e.g. 2015-03-21).',
    'admission_number': 'Optional.',
    'gender': 'Optional. One of: male, female, other.',
    'parent_name': 'Optional.',
    'parent_phone': 'Optional.',
}

# A column added to one tuple and forgotten in another would mean a template the importer
# rejects, or a column silently dropped. Catch that here rather than in a support ticket.
assert set(ALL_COLUMNS) == set(REQUIRED_COLUMNS) | set(OPTIONAL_COLUMNS)
assert set(COLUMN_HELP) == set(ALL_COLUMNS)

# Illustrative rows, not real people. Two rows so the shape of a filled sheet is obvious.
SAMPLE_ROWS: List[Dict[str, str]] = [
    {
        'email': 'arun.kumar@example.com',
        'full_name': 'Arun Kumar',
        'roll_number': '1',
        'date_of_birth': '2015-03-21',
        'admission_number': 'ADM1001',
        'gender': 'male',
        'parent_name': 'Ravi Kumar',
        'parent_phone': '9876543210',
    },
    {
        # Deliberately email-less and DOB-less: the clearest way to show those columns may
        # be left blank is a sample row that leaves them blank.
        'email': '',
        'full_name': 'Divya Raj',
        'roll_number': '2',
        'date_of_birth': '',
        'admission_number': 'ADM1002',
        'gender': 'female',
        'parent_name': 'Suresh Raj',
        'parent_phone': '9876543211',
    },
]


def build_csv_template() -> bytes:
    """The template as CSV bytes: a header row plus the sample rows."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(ALL_COLUMNS), lineterminator='\n')
    writer.writeheader()
    writer.writerows(SAMPLE_ROWS)
    # utf-8-sig: Excel opens a plain utf-8 CSV as mojibake for non-ASCII names.
    return buffer.getvalue().encode('utf-8-sig')


def build_xlsx_template() -> bytes:
    """The template as an .xlsx workbook: a Students sheet plus an Instructions sheet."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    header_font = Font(bold=True, color='FFFFFF')
    required_fill = PatternFill('solid', fgColor='2563EB')
    optional_fill = PatternFill('solid', fgColor='64748B')

    ws.append(list(ALL_COLUMNS))
    for idx, column in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = required_fill if column in REQUIRED_COLUMNS else optional_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(column) + 4, 18)

    for row in SAMPLE_ROWS:
        ws.append([row.get(column, '') for column in ALL_COLUMNS])

    ws.freeze_panes = 'A2'

    notes = wb.create_sheet('Instructions')
    notes.column_dimensions['A'].width = 24
    notes.column_dimensions['B'].width = 60
    notes.append(['Column', 'Notes'])
    notes.cell(row=1, column=1).font = Font(bold=True)
    notes.cell(row=1, column=2).font = Font(bold=True)
    for column in ALL_COLUMNS:
        notes.append([column, COLUMN_HELP[column]])
    notes.append([])
    notes.append(['Class / Section', 'Chosen on the upload form — not columns in this file.'])
    notes.append(['Sample rows', 'Delete the two example rows before uploading.'])
    notes.append(['Headings', 'Keep row 1 as-is. Spacing and capitalisation are forgiving.'])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
